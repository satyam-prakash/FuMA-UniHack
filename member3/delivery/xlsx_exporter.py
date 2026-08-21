"""
XLSX writer for the 252-column delivery format.

Owned by Member 3. The client reviews the delivery file by hand, so the header
row is bold, frozen and filterable across all 252 columns.
"""

import io
from typing import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from member3.delivery.columns import DELIVERY_COLUMN_COUNT, DELIVERY_COLUMNS
from member3.delivery.validators import validate_delivery_rows

SHEET_NAME = "Delivery Format"


def rows_to_xlsx_bytes(rows: Sequence[Mapping[str, str]]) -> bytes:
    """Serialises delivery rows to a single-sheet XLSX workbook as bytes."""
    validate_delivery_rows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    sheet.append(DELIVERY_COLUMNS)
    bold = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold

    for row in rows:
        sheet.append([row[column] for column in DELIVERY_COLUMNS])

    sheet.freeze_panes = "A2"
    last_column = get_column_letter(DELIVERY_COLUMN_COUNT)
    sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
