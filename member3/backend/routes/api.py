"""
FuMA Member 3 API routes.

One router module: the endpoint set is small and entirely about job lifecycle,
so splitting it across seven files would add navigation cost without removing
any duplication.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, Response

from member3.backend.models.api_models import EnrichRequest, ReviewRequest
from member3.backend.services.batch_service import get_job_store
from member3.backend.services.metrics_service import compute_benchmark, compute_metrics
from member3.backend.services.pipeline_service import (
    REQUIRED_INPUT_COLUMNS,
    validate_input_columns,
)
from member3.delivery.columns import DELIVERY_COLUMN_COUNT
from member3.delivery.csv_exporter import rows_to_csv_bytes
from member3.delivery.mapper import map_record_to_delivery
from member3.delivery.validators import DeliveryValidationError, check_delivery
from member3.delivery.xlsx_exporter import rows_to_xlsx_bytes

router = APIRouter(prefix="/api")

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
SAMPLE_CSV = DATA_DIR / "sample_input_1000.csv"
GROUND_TRUTH_CSV = DATA_DIR / "expected_delivery_format.csv"

MAX_UPLOAD_BYTES = 8 * 1024 * 1024
ALLOWED_SUFFIXES = {".csv", ".xlsx"}


def _error(status: int, code: str, message: str, details: List[str] | None = None) -> JSONResponse:
    return JSONResponse(
        status_code=status,
        content={"error": {"code": code, "message": message, "row_id": None, "details": details or []}},
    )


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]", "_", Path(name or "upload.csv").name)
    return cleaned or "upload.csv"


def _parse_csv_bytes(payload: bytes) -> tuple[List[Dict[str, Any]], List[str]]:
    text = payload.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = [{k: (v if v is not None else "") for k, v in row.items() if k is not None} for row in reader]
    return rows, list(reader.fieldnames or [])


def _parse_xlsx_bytes(payload: bytes) -> tuple[List[Dict[str, Any]], List[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    header = ["" if cell is None else str(cell) for cell in next(iterator, [])]
    rows = []
    for values in iterator:
        if values is None or all(v is None for v in values):
            continue
        rows.append({header[i]: ("" if v is None else str(v)) for i, v in enumerate(values) if i < len(header)})
    return rows, header


def _duplicate_headers(header: Sequence[str]) -> List[str]:
    """Case-insensitive duplicate column names. csv.DictReader silently keeps
    the last value for a duplicated key, which would hide data corruption, so
    duplicates are rejected before any row is built."""
    seen: Dict[str, int] = {}
    dupes: List[str] = []
    for name in header or []:
        key = str(name).strip().lower()
        if not key:
            continue
        seen[key] = seen.get(key, 0) + 1
    for key, count in seen.items():
        if count > 1:
            dupes.append(key)
    return dupes


def _upload_response(summary: Dict[str, Any]) -> Dict[str, Any]:
    """Job summaries expose ``total``; the upload contract calls it ``rows``."""
    return {
        "job_id": summary["job_id"],
        "filename": summary["filename"],
        "rows": summary["total"],
        "columns": list(REQUIRED_INPUT_COLUMNS),
        "status": summary["status"],
    }


def _require_job(job_id: str) -> Optional[Dict[str, Any]]:
    """Returns the job or None. Callers return ``_job_not_found(job_id)``."""
    return get_job_store().get(job_id)


def _job_not_found(job_id: str) -> JSONResponse:
    return _error(404, "JOB_NOT_FOUND", f"Unknown job_id {job_id}")


def _ground_truth_rows() -> List[Dict[str, Any]]:
    if not GROUND_TRUTH_CSV.exists():
        return []
    with GROUND_TRUTH_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def _deliverable_results(job: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Rows eligible for delivery: anything that produced an enriched record."""
    return [r for r in job["results"] if r.get("enriched")]


def _delivery_rows(job: Dict[str, Any]) -> tuple[List[Dict[str, str]], List[str]]:
    """Maps a job's enriched results into delivery rows plus mapping warnings."""
    rows: List[Dict[str, str]] = []
    warnings: List[str] = []
    for result in _deliverable_results(job):
        row, row_warnings = map_record_to_delivery(result["enriched"], result.get("raw") or {})
        rows.append(row)
        warnings.extend(row_warnings)
    return rows, warnings


# --------------------------------------------------------------------- health


@router.get("/health")
def health() -> Dict[str, Any]:
    return {"status": "ok", "service": "fuma-api", "delivery_columns": DELIVERY_COLUMN_COUNT}


# --------------------------------------------------------------------- upload


@router.post("/upload")
async def upload(file: UploadFile = File(...)):
    filename = _safe_filename(file.filename or "")
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        return _error(400, "UNSUPPORTED_FILE_TYPE", f"Only {sorted(ALLOWED_SUFFIXES)} are accepted", [filename])

    payload = await file.read()
    if not payload:
        return _error(400, "EMPTY_FILE", "Uploaded file is empty")
    if len(payload) > MAX_UPLOAD_BYTES:
        return _error(413, "FILE_TOO_LARGE", f"Upload exceeds {MAX_UPLOAD_BYTES} bytes")

    try:
        rows, header = _parse_xlsx_bytes(payload) if suffix == ".xlsx" else _parse_csv_bytes(payload)
    except Exception as exc:
        return _error(400, "PARSE_ERROR", "Could not parse uploaded file", [str(exc)])

    dupes = _duplicate_headers(header)
    if dupes:
        return _error(
            400,
            "DUPLICATE_COLUMNS",
            "Uploaded file contains duplicate column names",
            [f"duplicates: {', '.join(dupes)}"],
        )

    if not rows:
        return _error(400, "NO_DATA_ROWS", "File contains a header but no data rows")

    missing = validate_input_columns(header)
    if missing:
        return _error(
            422,
            "MISSING_REQUIRED_COLUMNS",
            "Input is missing required columns",
            [f"missing: {', '.join(missing)}", f"required: {', '.join(REQUIRED_INPUT_COLUMNS)}"],
        )

    return _upload_response(get_job_store().create_job(filename, rows))


@router.post("/demo/sample")
def load_sample():
    if not SAMPLE_CSV.exists():
        return _error(500, "SAMPLE_MISSING", "Bundled sample dataset not found")
    rows, _ = _parse_csv_bytes(SAMPLE_CSV.read_bytes())
    return _upload_response(get_job_store().create_job(SAMPLE_CSV.name, rows))


# -------------------------------------------------------------------- enrich


@router.post("/enrich")
def enrich(request: EnrichRequest):
    store = get_job_store()
    job = _require_job(request.job_id)
    if job is None:
        return _job_not_found(request.job_id)
    if job["status"] in ("queued", "processing"):
        return {"job_id": job["job_id"], "status": job["status"]}
    summary = store.start_job_async(request.job_id)
    return {"job_id": summary["job_id"], "status": summary["status"]}


# ---------------------------------------------------------------------- jobs


@router.get("/jobs/{job_id}")
def job_status(job_id: str):
    if _require_job(job_id) is None:
        return _job_not_found(job_id)
    return get_job_store().summary(job_id)


@router.get("/jobs/{job_id}/results")
def job_results(
    job_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    status: str = "all",
    search: str = "",
):
    if _require_job(job_id) is None:
        return _job_not_found(job_id)
    return get_job_store().list_results(job_id, page, page_size, status, search)


@router.get("/jobs/{job_id}/results/{row_id}")
def job_result(job_id: str, row_id: int):
    if _require_job(job_id) is None:
        return _job_not_found(job_id)
    result = get_job_store().get_result(job_id, row_id)
    if result is None:
        return _error(404, "ROW_NOT_FOUND", f"Unknown row_id {row_id}")

    payload = dict(result)
    mpn = str((result.get("enriched") or {}).get("mfg_part_num", "")).strip().upper()
    payload["ground_truth"] = next(
        (
            row
            for row in _ground_truth_rows()
            if str(row.get("Mfg_Part_Num", "")).strip().upper() == mpn and mpn
        ),
        None,
    )
    return payload


@router.get("/jobs/{job_id}/metrics")
def job_metrics(job_id: str):
    job = _require_job(job_id)
    if job is None:
        return _job_not_found(job_id)
    metrics = compute_metrics(job["results"])
    metrics["benchmark"] = compute_benchmark(job["results"], _ground_truth_rows())
    metrics["elapsed_seconds"] = job.get("elapsed_seconds", 0.0)
    metrics["delivery_columns"] = DELIVERY_COLUMN_COUNT
    return metrics


# -------------------------------------------------------------------- review


@router.get("/jobs/{job_id}/review")
def review_queue(job_id: str):
    if _require_job(job_id) is None:
        return _job_not_found(job_id)
    return {"rows": get_job_store().review_rows(job_id) or []}


@router.post("/jobs/{job_id}/review/{row_id}")
def submit_review(job_id: str, row_id: int, request: ReviewRequest):
    if _require_job(job_id) is None:
        return _job_not_found(job_id)
    result = get_job_store().set_review(job_id, row_id, request.action, request.comment)
    if result is None:
        return _error(404, "ROW_NOT_FOUND", f"Unknown row_id {row_id}")
    return {"row_id": row_id, "review": result["review"]}


# -------------------------------------------------------------------- export


@router.get("/jobs/{job_id}/export/status")
def export_status(job_id: str):
    job = _require_job(job_id)
    if job is None:
        return _job_not_found(job_id)
    rows, warnings = _delivery_rows(job)
    valid, errors = check_delivery(rows)
    return {
        "delivery_columns": DELIVERY_COLUMN_COUNT,
        "valid": valid,
        "errors": errors + warnings,
        "row_count": len(rows),
        "rows_needing_review": sum(1 for r in job["results"] if r["review"]["needs_review"]),
    }


def _validated_delivery_rows(job: Dict[str, Any]) -> Optional[List[Dict[str, str]]]:
    rows, _ = _delivery_rows(job)
    valid, errors = check_delivery(rows)
    if not valid:
        raise DeliveryValidationError("; ".join(errors))
    return rows


@router.get("/jobs/{job_id}/export.csv")
def export_job_csv(job_id: str):
    job = _require_job(job_id)
    if job is None:
        return _job_not_found(job_id)
    try:
        payload = rows_to_csv_bytes(_validated_delivery_rows(job))
    except DeliveryValidationError as exc:
        return _error(409, "DELIVERY_VALIDATION_FAILED", "Delivery validation failed", [str(exc)])
    job["status"] = "exported"
    return Response(
        content=payload,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="fuma_delivery_{job_id}.csv"'},
    )


@router.get("/jobs/{job_id}/export.xlsx")
def export_job_xlsx(job_id: str):
    job = _require_job(job_id)
    if job is None:
        return _job_not_found(job_id)
    try:
        payload = rows_to_xlsx_bytes(_validated_delivery_rows(job))
    except DeliveryValidationError as exc:
        return _error(409, "DELIVERY_VALIDATION_FAILED", "Delivery validation failed", [str(exc)])
    job["status"] = "exported"
    return Response(
        content=payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="fuma_delivery_{job_id}.xlsx"'},
    )
