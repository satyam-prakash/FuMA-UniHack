"""Tests for the Member 3 delivery export layer (252-column contract)."""

import csv
import io
from pathlib import Path

import openpyxl
import pytest

from member3.delivery.columns import (
    ATTRIBUTE_SLOTS,
    DELIVERY_COLUMN_COUNT,
    DELIVERY_COLUMNS,
)
from member3.delivery.csv_exporter import rows_to_csv_bytes
from member3.delivery.mapper import FEATURE_SLOTS, map_record_to_delivery
from member3.delivery.validators import (
    DeliveryValidationError,
    check_delivery,
    validate_delivery_headers,
    validate_delivery_rows,
)
from member3.delivery.xlsx_exporter import SHEET_NAME, rows_to_xlsx_bytes

DATA_DIR = Path(__file__).resolve().parents[1] / "data"
EXPECTED_CSV = DATA_DIR / "expected_delivery_format.csv"
SAMPLE_CSV = DATA_DIR / "sample_input_1000.csv"

RAW = {
    "Mfg_Part_Num": "PDSH4816AF",
    "Part_Desc": "PDSH4816AF Dishwasher SS - Display Only",
    "E1_Brand": "-- Unbranded --",
    "Unilog_Brand": "-- No Unilog Brand --",
    "DIB_Brand": "-- No DIB Brand --",
    "Part_Manuf": "Appliance Dealers Cooperative (APPDE)",
}


def enriched(**overrides):
    record = {
        "mfg_part_num": "PDSH4816AF",
        "part_desc_raw": RAW["Part_Desc"],
        "manufacturer_name": "Rheem Manufacturing",
        "brand_name": "FRIGIDAIRE®",
        "series": "",
        "classpath": "Appliances & Consumer Electronics>Kitchen Appliances>Built-In Dishwashers",
        "unspsc": "52141505",
        "product_name": "Dishwasher",
        "attributes": [],
        "features": [],
        "invoice_desc": "DISHWASHER SST",
        "mobile_desc": "Rheem Manufacturing FRIGIDAIRE, Dishwasher, PDSH4816AF",
        "short_desc": "FRIGIDAIRE® PDSH4816AF Dishwasher, Stainless Steel",
        "long_desc1": "FRIGIDAIRE® Dishwasher, Stainless Steel",
        "retail_desc": "Dishwasher, Stainless Steel",
        "confidence_score": 100.0,
        "needs_review": False,
        "review_reasons": [],
    }
    record.update(overrides)
    return record


def attrs(count):
    return [
        {"label": f"Attr {i}", "value": f"Val {i}", "uom": "in"} for i in range(1, count + 1)
    ]


# --- columns contract -------------------------------------------------------


def test_column_count():
    assert len(DELIVERY_COLUMNS) == 252
    assert DELIVERY_COLUMN_COUNT == 252


def test_columns_match_reference_csv_header():
    with EXPECTED_CSV.open(encoding="utf-8-sig", newline="") as handle:
        header = next(csv.reader(handle))
    assert header == DELIVERY_COLUMNS


def test_no_duplicate_headers():
    assert len(set(DELIVERY_COLUMNS)) == len(DELIVERY_COLUMNS)


# --- mapper ----------------------------------------------------------------


def test_mapped_row_is_exactly_252_keys_in_order():
    row, warnings = map_record_to_delivery(enriched(), RAW)
    assert list(row) == DELIVERY_COLUMNS
    assert warnings == []
    assert all(isinstance(value, str) for value in row.values())


def test_passthrough_and_enriched_fields():
    row, _ = map_record_to_delivery(enriched(), RAW)
    assert row["Mfg_Part_Num"] == "PDSH4816AF"
    assert row["Part_Desc"] == RAW["Part_Desc"]
    assert row["E1_Brand"] == RAW["E1_Brand"]
    assert row["Part_Manuf"] == RAW["Part_Manuf"]
    assert row["MANUFACTURER_NAME"] == "Rheem Manufacturing"
    assert row["BRAND_NAME"] == "FRIGIDAIRE®"
    assert row["MANUFACTURER_PART_NUMBER"] == "PDSH4816AF"
    assert row["Product Name"] == "Dishwasher"
    assert row["UNSPSC"] == "52141505"
    # Absent raw keys and out-of-scope columns stay blank, never invented.
    assert row["PART_NUMBER"] == ""
    assert row["Dept"] == ""
    assert row["MFR URL"] == ""
    assert row["Ref URL 5"] == ""
    assert row["TRADE_NAME"] == ""
    assert row["MARKETING_DESCRIPTION"] == ""
    assert row["Country Of Origin"] == ""


def test_none_and_nan_become_blank():
    row, _ = map_record_to_delivery(
        enriched(unspsc=None, series=None), {**RAW, "PART_NUMBER": float("nan")}
    )
    assert row["UNSPSC"] == ""
    assert row["PART_NUMBER"] == ""


def test_attribute_slot_mapping():
    row, warnings = map_record_to_delivery(enriched(attributes=attrs(3)), RAW)
    assert warnings == []
    for slot in (1, 2, 3):
        assert row[f"ATTRIBUTE_LABEL {slot}"] == f"Attr {slot}"
        assert row[f"ATTRIBUTE_VALUE {slot}"] == f"Val {slot}"
        assert row[f"ATTRIBUTE_UOM {slot}"] == "in"
    for slot in range(4, ATTRIBUTE_SLOTS + 1):
        assert row[f"ATTRIBUTE_LABEL {slot}"] == ""
        assert row[f"ATTRIBUTE_VALUE {slot}"] == ""
        assert row[f"ATTRIBUTE_UOM {slot}"] == ""


def test_attribute_overflow_keeps_252_columns_and_warns():
    row, warnings = map_record_to_delivery(enriched(attributes=attrs(60)), RAW)
    assert len(row) == 252
    assert list(row) == DELIVERY_COLUMNS
    assert warnings == [
        "attribute overflow: 60 attributes extracted, only first 50 exported"
    ]
    assert row["ATTRIBUTE_LABEL 50"] == "Attr 50"
    assert row["ATTRIBUTE_VALUE 50"] == "Val 50"
    assert "ATTRIBUTE_LABEL 51" not in row


def test_feature_overflow_keeps_20_slots_and_warns():
    features = [f"Feature {i}" for i in range(1, 26)]
    row, warnings = map_record_to_delivery(enriched(features=features), RAW)
    assert len(row) == 252
    assert warnings == ["feature overflow: 25 features extracted, only first 20 exported"]
    assert row["ITEM_FEATURES_1"] == "Feature 1"
    assert row[f"ITEM_FEATURES_{FEATURE_SLOTS}"] == "Feature 20"
    assert "ITEM_FEATURES_21" not in row


def test_features_under_limit_fill_in_order():
    row, warnings = map_record_to_delivery(enriched(features=["A", "B"]), RAW)
    assert warnings == []
    assert row["ITEM_FEATURES_1"] == "A"
    assert row["ITEM_FEATURES_2"] == "B"
    assert row["ITEM_FEATURES_3"] == ""


# --- validators ------------------------------------------------------------


def test_validate_headers_accepts_contract():
    validate_delivery_headers(DELIVERY_COLUMNS)
    validate_delivery_headers(tuple(DELIVERY_COLUMNS))


@pytest.mark.parametrize(
    "headers",
    [
        [DELIVERY_COLUMNS[1], DELIVERY_COLUMNS[0]] + DELIVERY_COLUMNS[2:],  # reordered
        DELIVERY_COLUMNS[:-1],  # short
        DELIVERY_COLUMNS + ["EXTRA"],  # extended
    ],
    ids=["reordered", "short", "extended"],
)
def test_validate_headers_rejects_bad_header_lists(headers):
    with pytest.raises(DeliveryValidationError):
        validate_delivery_headers(headers)


def test_validate_rows_accepts_mapped_row():
    row, _ = map_record_to_delivery(enriched(), RAW)
    validate_delivery_rows([row])


def test_validate_rows_rejects_missing_key():
    row, _ = map_record_to_delivery(enriched(), RAW)
    del row["UNSPSC"]
    with pytest.raises(DeliveryValidationError, match="missing columns"):
        validate_delivery_rows([row])


def test_validate_rows_rejects_extra_key():
    row, _ = map_record_to_delivery(enriched(), RAW)
    row["BONUS_COLUMN"] = "x"
    with pytest.raises(DeliveryValidationError, match="unexpected columns"):
        validate_delivery_rows([row])


def test_check_delivery_reports_instead_of_raising():
    good, _ = map_record_to_delivery(enriched(), RAW)
    assert check_delivery([good]) == (True, [])

    missing, _ = map_record_to_delivery(enriched(), RAW)
    del missing["UNSPSC"]
    ok, errors = check_delivery([missing])
    assert ok is False
    assert errors and "missing columns" in errors[0]

    extra, _ = map_record_to_delivery(enriched(), RAW)
    extra["BONUS_COLUMN"] = "x"
    ok, errors = check_delivery([extra])
    assert ok is False
    assert errors and "unexpected columns" in errors[0]


# --- csv exporter ----------------------------------------------------------


def test_csv_round_trip_preserves_registered_symbol():
    row, _ = map_record_to_delivery(enriched(attributes=attrs(2)), RAW)
    data = rows_to_csv_bytes([row])

    assert data.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM so Excel keeps ® and ™
    reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
    assert reader.fieldnames == DELIVERY_COLUMNS
    assert len(reader.fieldnames) == 252

    parsed = list(reader)
    assert len(parsed) == 1
    assert parsed[0]["BRAND_NAME"] == "FRIGIDAIRE®"
    assert "®" in parsed[0]["SHORT_DESC"]
    assert dict(parsed[0]) == row


def test_csv_uses_crlf_and_validates_rows():
    row, _ = map_record_to_delivery(enriched(), RAW)
    assert b"\r\n" in rows_to_csv_bytes([row])

    del row["UNSPSC"]
    with pytest.raises(DeliveryValidationError):
        rows_to_csv_bytes([row])


# --- xlsx exporter ---------------------------------------------------------


def test_xlsx_shape_and_formatting():
    row, _ = map_record_to_delivery(enriched(attributes=attrs(2)), RAW)
    workbook = openpyxl.load_workbook(io.BytesIO(rows_to_xlsx_bytes([row, row])))
    sheet = workbook.active

    assert sheet.title == SHEET_NAME
    assert sheet.max_column == 252
    assert sheet.max_row == 3
    assert [cell.value for cell in sheet[1]] == DELIVERY_COLUMNS
    assert all(cell.font.bold for cell in sheet[1])
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:IR3"
    assert sheet.cell(row=2, column=DELIVERY_COLUMNS.index("BRAND_NAME") + 1).value == (
        "FRIGIDAIRE®"
    )


def test_xlsx_validates_rows():
    row, _ = map_record_to_delivery(enriched(), RAW)
    row["BONUS_COLUMN"] = "x"
    with pytest.raises(DeliveryValidationError):
        rows_to_xlsx_bytes([row])


# --- end to end ------------------------------------------------------------


def test_end_to_end_real_row_through_member1_and_member2():
    from fuma_engine.pipeline_interface import enrich_single_item
    from fuma_rules.stage1_master_data import MasterDataPipelineStage

    with SAMPLE_CSV.open(encoding="utf-8-sig", newline="") as handle:
        raw = next(r for r in csv.DictReader(handle) if r["Mfg_Part_Num"] == "PDSH4816AF")

    normalized = MasterDataPipelineStage().process_item(raw)
    record = enrich_single_item(normalized)
    row, warnings = map_record_to_delivery(record, raw)

    assert list(row) == DELIVERY_COLUMNS
    assert len(row) == 252
    assert warnings == []
    assert row["Mfg_Part_Num"] == "PDSH4816AF"  # raw passthrough survived
    assert row["Part_Desc"] == raw["Part_Desc"]
    assert row["MANUFACTURER_PART_NUMBER"] == record["mfg_part_num"]
    assert row["Classpath"] == record["classpath"]
    assert row["Product Name"] == record["product_name"]
    assert row["INVOICE_DESC"] == record["invoice_desc"]

    validate_delivery_rows([row])
    reader = csv.DictReader(io.StringIO(rows_to_csv_bytes([row]).decode("utf-8-sig")))
    assert dict(next(iter(reader))) == row
