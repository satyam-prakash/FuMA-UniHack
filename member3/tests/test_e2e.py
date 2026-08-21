"""
End-to-end verification of the Member 3 stack.

Drives the real FastAPI app through the full demo path: upload the bundled
1,000-row dataset, enrich it, read metrics, download both delivery files and
re-parse them to prove the 252-column contract holds.
"""

from __future__ import annotations

import csv
import io
import time

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from member3.backend.main import app
from member3.delivery.columns import DELIVERY_COLUMNS

SAMPLE = "member3/data/sample_input_1000.csv"


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as test_client:
        yield test_client


def _wait_for_completion(client, job_id: str, timeout: float = 180.0) -> dict:
    deadline = time.time() + timeout
    while time.time() < deadline:
        payload = client.get(f"/api/jobs/{job_id}").json()
        if payload["status"] in ("completed", "completed_with_review", "failed", "exported"):
            return payload
        time.sleep(0.15)
    raise AssertionError(f"job {job_id} did not finish within {timeout}s")


def test_health(client):
    body = client.get("/api/health").json()
    assert body["status"] == "ok"
    assert body["delivery_columns"] == 252


def test_upload_rejects_bad_extension(client):
    response = client.post("/api/upload", files={"file": ("notes.txt", b"a,b\n1,2\n", "text/plain")})
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "UNSUPPORTED_FILE_TYPE"


def test_upload_rejects_empty_file(client):
    response = client.post("/api/upload", files={"file": ("empty.csv", b"", "text/csv")})
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "EMPTY_FILE"


def test_upload_rejects_missing_columns(client):
    payload = b"Mfg_Part_Num,Part_Desc\nABC,Widget\n"
    response = client.post("/api/upload", files={"file": ("partial.csv", payload, "text/csv")})
    assert response.status_code == 422
    body = response.json()["error"]
    assert body["code"] == "MISSING_REQUIRED_COLUMNS"
    assert "E1_Brand" in " ".join(body["details"])


def test_upload_rejects_duplicate_headers(client):
    payload = b"Mfg_Part_Num,Mfg_Part_Num,Part_Desc,E1_Brand,Unilog_Brand,DIB_Brand,Part_Manuf\nA,B,C,D,E,F,G\n"
    response = client.post("/api/upload", files={"file": ("dup.csv", payload, "text/csv")})
    assert response.status_code == 400
    body = response.json()["error"]
    assert body["code"] == "DUPLICATE_COLUMNS"
    assert "mfg_part_num" in " ".join(body["details"]).lower()


def test_upload_rejects_duplicate_optional_headers(client):
    payload = b"Mfg_Part_Num,Part_Desc,E1_Brand,Unilog_Brand,DIB_Brand,Part_Manuf,Part_Manuf\nA,B,C,D,E,F,G\n"
    response = client.post("/api/upload", files={"file": ("dup2.csv", payload, "text/csv")})
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "DUPLICATE_COLUMNS"


def test_unknown_job_is_404_with_envelope(client):
    response = client.get("/api/jobs/job_does_not_exist")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "JOB_NOT_FOUND"


def test_unknown_api_route_is_json_404(client):
    for path in ("/api/does-not-exist", "/api/foo/bar", "/api/jobs/does-not-exist/export.pdf"):
        response = client.get(path)
        assert response.status_code == 404
        body = response.json()
        assert "error" in body, f"{path} must return the JSON error envelope"
        assert body["error"]["code"] == "NOT_FOUND"


def test_unknown_row_is_404_with_envelope(client):
    upload = client.post("/api/demo/sample").json()
    job_id = upload["job_id"]
    assert client.post("/api/enrich", json={"job_id": job_id}).status_code == 200
    _wait_for_completion(client, job_id)
    response = client.get(f"/api/jobs/{job_id}/results/99999")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "ROW_NOT_FOUND"


def test_non_integer_row_id_uses_error_envelope(client):
    upload = client.post("/api/demo/sample").json()
    response = client.get(f"/api/jobs/{upload['job_id']}/results/abc")
    assert response.status_code == 422
    assert response.json()["error"]["code"] == "VALIDATION_ERROR"


def test_benchmark_is_none_when_nothing_matches(client):
    payload = (
        b"Mfg_Part_Num,Part_Desc,E1_Brand,Unilog_Brand,DIB_Brand,Part_Manuf\n"
        b"NO-SUCH-MPN,Widget,--,--,--,Acme (ACME)\n"
    )
    upload = client.post("/api/upload", files={"file": ("nm.csv", payload, "text/csv")}).json()
    job_id = upload["job_id"]
    client.post("/api/enrich", json={"job_id": job_id})
    _wait_for_completion(client, job_id)
    metrics = client.get(f"/api/jobs/{job_id}/metrics").json()
    assert metrics["benchmark"] is None


def test_full_demo_path(client):
    upload = client.post("/api/demo/sample").json()
    job_id = upload["job_id"]
    assert upload["rows"] == 1000, upload

    assert client.post("/api/enrich", json={"job_id": job_id, "mode": "demo"}).status_code == 200
    status = _wait_for_completion(client, job_id)

    assert status["processed"] == status["total"]
    assert status["success"] + status["review"] + status["errors"] == status["total"]
    assert status["progress"] == 100

    metrics = client.get(f"/api/jobs/{job_id}/metrics").json()
    assert metrics["total"] == status["total"]
    assert metrics["invoice_char_pass"] == 100.0, "every invoice description must fit 40 chars"
    assert metrics["invoice_caps_pass"] == 100.0, "every invoice description must be uppercase"
    assert metrics["schema_pass_rate"] == 100.0, "no ProductRecord schema failures allowed"
    assert "mobile_target_60_80_pass" in metrics and "schema_mobile_pass" in metrics
    assert metrics["benchmark"]["matched_rows"] >= 2

    page = client.get(f"/api/jobs/{job_id}/results", params={"page": 1, "page_size": 25}).json()
    assert len(page["rows"]) == 25
    assert page["total"] == status["total"]
    assert page["job_id"] == job_id
    assert page["pages"] == 40

    detail = client.get(f"/api/jobs/{job_id}/results/1").json()
    assert detail["row_id"] == 1
    assert detail["enriched"]["invoice_desc"]

    searched = client.get(f"/api/jobs/{job_id}/results", params={"search": "PDSH4816AF"}).json()
    assert searched["total"] >= 1

    review = client.get(f"/api/jobs/{job_id}/review").json()
    if review["rows"]:
        entry = review["rows"][0]
        # Compact queue entries: no 252-column delivery payloads in the queue.
        assert {"row_id", "mpn", "confidence_score", "status", "reasons", "categories", "decision"} <= set(entry)
        row_id = entry["row_id"]
        decision = client.post(
            f"/api/jobs/{job_id}/review/{row_id}",
            json={"action": "approve", "comment": "verified against source"},
        ).json()
        assert decision["review"]["decision"]["action"] == "approve"
        assert decision["review"]["decision"]["comment"] == "verified against source"
        assert "at" in decision["review"]["decision"]
        assert decision["review"]["needs_review"] is False

    export_state = client.get(f"/api/jobs/{job_id}/export/status").json()
    assert export_state["valid"] is True
    assert export_state["delivery_columns"] == 252

    csv_response = client.get(f"/api/jobs/{job_id}/export.csv")
    assert csv_response.status_code == 200
    text = csv_response.content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    header = next(reader)
    data_rows = [r for r in reader if r]
    assert header == DELIVERY_COLUMNS
    assert len(header) == 252
    assert len(data_rows) == export_state["row_count"]
    assert all(len(r) == 252 for r in data_rows)
    assert "®" in text

    xlsx_response = client.get(f"/api/jobs/{job_id}/export.xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    xlsx_header = ["" if c is None else str(c) for c in next(rows)]
    assert xlsx_header == DELIVERY_COLUMNS
    assert sum(1 for r in rows if r and any(v is not None for v in r)) == export_state["row_count"]
